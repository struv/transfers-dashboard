"""Shared aggregation logic used by both build.py (2026) and
validate_2024.py.

Encapsulates the rules described in PLAN.md so a single source of truth
defines what counts as Transferred / Total / Adults / Peds / SHI / Other.
"""
from __future__ import annotations

import datetime as _dt
import re
from dataclasses import dataclass, field
from typing import Iterable

import openpyxl


# --- Sheet classification ---------------------------------------------------

# Substring rules applied in priority order. First match wins.
# BEHRO covers BEHROOZAN, BEHROO, "BENJAMIN BEHRO".
GROUP_RULES = [
    ("hanna",     "Hanna",     "Dr. Hanna acquired IPAs",     ("HANNA",)),
    ("samala",    "Samala",    "Dr. Samala acquired IPAs",    ("SAMALA",)),
    ("benny_b",   "Benny B",   "Dr. Behroozan acquired IPAs", ("BEHRO",)),
    ("sakhai",    "Sakhai",    "Dr. Sakhai acquired IPAs",    ("SAKHAI",)),
    ("la_mirada", "La Mirada", "Original La Mirada group",    ("LA MIRADA",)),
]

# Explicit assignments for sheets whose names don't contain any classifier
# substring. Keys are uppercase, whitespace-stripped sheet names.
SHEET_OVERRIDES = {
    "HOLLYWOOD PRESBYTERIAN GLOBAL":     "benny_b",
    "SOUTHERN CALIFORNIA ALTA GLOBAL":   "benny_b",
    "OPTUM ARTA MONARCH SOUTH COAST":    "la_mirada",
}

# Sheets that are shared / non-IPA-specific and should never appear in a
# group section.
EXCLUDED_SHEETS = {
    "DAILY LIST",
    "DAILY LIST 2",
    "DAILY LIST SLOW DO NOT USE",
    "HEALTH CARE OPTIONS",
    "HEALTHCARE OPTIONS",
    "CHDP",
    "LA CARE SHI DRS",
    "LA CARE SHI CONTRACTED DOCTORS",
    "MOLINA SHI DRS",
    "MOLINA SHI CONTRACTED DOCTORS",
    "REGAL DRS",
    "REGAL CONTRACTED DOCTORS",
    "DRS COMMERCIAL",
    "OFFICE EXTENSION",
    "MA LIST",
    "DETAIL1",
    "DETAIL2",
    "DETAIL3",
    "TOTALS",
}


def _group_meta(key: str) -> tuple[str, str, str]:
    for k, display, subtitle, _ in GROUP_RULES:
        if k == key:
            return (k, display, subtitle)
    raise KeyError(key)


def classify_sheet(name: str) -> tuple[str, str, str] | None:
    """Return (key, display_name, subtitle) or None if excluded."""
    n = " ".join(name.strip().upper().split())  # collapse whitespace
    if n in EXCLUDED_SHEETS:
        return None
    if n in SHEET_OVERRIDES:
        return _group_meta(SHEET_OVERRIDES[n])
    for key, display, subtitle, needles in GROUP_RULES:
        for needle in needles:
            if needle in n:
                return (key, display, subtitle)
    return None


# --- IPA pretty name --------------------------------------------------------

_SUFFIXES = [
    " DR BEHROOZAN", " BENJAMIN BEHRO", " BEHROOZAN", " DR BEHROO",
    " DR SAKHAI, YUSSEF", " DR SAKHAI", " DR. SAKHAI",
    "-DR HANNA", "-HANNA",
    "-SAMALA MD", "-SAMALA",
    " LA MIRADA",
]


def pretty_ipa_name(sheet_name: str) -> str:
    n = sheet_name.strip()
    upper = n.upper()
    # Strip trailing classifier
    for suf in _SUFFIXES:
        if upper.endswith(suf):
            n = n[: len(n) - len(suf)]
            break
    n = n.strip().rstrip(",").strip()
    # Title-case-ish but preserve common acronyms.
    keep_upper = {"CFC", "PIH", "CHOC", "MD", "IPA", "SHI"}
    parts = []
    for tok in re.split(r"\s+", n):
        if not tok:
            continue
        if tok.upper() in keep_upper:
            parts.append(tok.upper())
        else:
            parts.append(tok.capitalize())
    return " ".join(parts) or sheet_name


# --- Header detection -------------------------------------------------------

REQUIRED_HEADERS = {
    "date":             ["date"],
    "dob":              ["dob"],
    "transfer_to":      ["transfer to"],
}

# Optional columns. Detected if present; absent → blank values.
OPTIONAL_HEADERS = {
    "transfer_status":  ["transfer status"],
    "effective_date":   ["effective date"],
    "patient_name":     ["patient name"],
    "insurance_carrier": ["insurance carrier"],
}


def _norm_header(v) -> str:
    if v is None:
        return ""
    return str(v).strip().lower().replace("\xa0", " ")


def detect_columns(header_row: tuple) -> dict[str, int] | None:
    cols: dict[str, int] = {}
    normalized = [_norm_header(c) for c in header_row]
    for key, needles in REQUIRED_HEADERS.items():
        for i, h in enumerate(normalized):
            if any(n == h for n in needles):
                cols[key] = i
                break
    if len(cols) != len(REQUIRED_HEADERS):
        return None
    for key, needles in OPTIONAL_HEADERS.items():
        for i, h in enumerate(normalized):
            if any(n == h for n in needles):
                cols[key] = i
                break
    return cols


# --- Date / DOB parsing -----------------------------------------------------

_DATE_PATTERNS = [
    "%m/%d/%Y",
    "%m/%d/%y",
    "%-m/%-d/%Y",
    "%Y-%m-%d",
    "%m-%d-%Y",
]


def coerce_date(v):
    """Return a datetime.date or None.

    Accepts:
      - datetime.datetime / datetime.date
      - numeric Excel serial (rare with data_only=True)
      - strings like ' 1/09/1981', '\\t1/9/81', '2024-11-01'
    """
    if v is None:
        return None
    if isinstance(v, _dt.datetime):
        return v.date()
    if isinstance(v, _dt.date):
        return v
    if isinstance(v, (int, float)):
        # Excel serial number; convert
        try:
            if v <= 0 or v > 80000:
                return None
            base = _dt.date(1899, 12, 30)
            return base + _dt.timedelta(days=int(v))
        except Exception:
            return None
    if isinstance(v, str):
        s = v.strip().replace("\t", "").replace("\xa0", " ").strip()
        if not s:
            return None
        for pat in _DATE_PATTERNS:
            try:
                return _dt.datetime.strptime(s, pat).date()
            except ValueError:
                continue
        # Try fuzzy m/d/y with single digits
        m = re.match(r"^(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})$", s)
        if m:
            mm, dd, yy = m.groups()
            yy_i = int(yy)
            if yy_i < 100:
                yy_i += 1900 if yy_i >= 30 else 2000
            try:
                return _dt.date(yy_i, int(mm), int(dd))
            except ValueError:
                return None
    return None


# --- Row classification -----------------------------------------------------

EXCLUDED_STATUS_PREFIXES = ("INACTIVE", "REFUSED", "NO PHONE", "LEFT MESSAGE")


def is_excluded_status(status) -> bool:
    if status is None:
        return False
    s = str(status).strip().upper().replace("\xa0", " ")
    if not s:
        return False
    for pref in EXCLUDED_STATUS_PREFIXES:
        if s.startswith(pref):
            return True
    return False


def is_shi(transfer_to) -> bool:
    if transfer_to is None:
        return False
    return "SHI" in str(transfer_to).upper()


def age_on(dob: _dt.date | None, ref: _dt.date | None) -> int | None:
    if dob is None or ref is None:
        return None
    years = ref.year - dob.year
    if (ref.month, ref.day) < (dob.month, dob.day):
        years -= 1
    if years < 0 or years > 130:
        return None
    return years


# --- Aggregates -------------------------------------------------------------

# Counters held inside each monthly bucket. Kept as a flat dict (instead of
# a nested IpaAgg) because buckets never recurse and the JSON stays compact.
MONTH_FIELDS: tuple[str, ...] = (
    "transferred", "total",
    "adults_t", "adults_total",
    "peds_t", "peds_total",
    "other_t", "shi_t",
    "adults_shi_t", "adults_other_t",
    "peds_shi_t", "peds_other_t",
    "unknown_age_t", "unknown_age_total",
)


def _new_month_bucket() -> dict[str, int]:
    return {k: 0 for k in MONTH_FIELDS}


def _merge_month_bucket(dst: dict[str, int], src: dict[str, int]) -> None:
    for k in MONTH_FIELDS:
        dst[k] = dst.get(k, 0) + src.get(k, 0)


@dataclass
class IpaAgg:
    ipa: str
    sheet: str
    transferred: int = 0
    total: int = 0
    adults_t: int = 0
    adults_total: int = 0
    peds_t: int = 0
    peds_total: int = 0
    other_t: int = 0
    shi_t: int = 0
    unknown_age_t: int = 0  # transferred rows with no usable DOB
    unknown_age_total: int = 0
    # Segment x channel cross counters (transferred only).
    # Unknown-age rows bucket into adults_* to preserve the identity
    # adults_*_t + peds_*_t == *_t.
    adults_shi_t: int = 0
    adults_other_t: int = 0
    peds_shi_t: int = 0
    peds_other_t: int = 0
    # Per-month buckets keyed YYYY-MM. Each value mirrors MONTH_FIELDS.
    by_month: dict[str, dict[str, int]] = field(default_factory=dict)

    def to_dict(self) -> dict:
        return {
            "ipa": self.ipa,
            "sheet": self.sheet,
            "transferred": self.transferred,
            "total": self.total,
            "adults_t": self.adults_t,
            "adults_total": self.adults_total,
            "peds_t": self.peds_t,
            "peds_total": self.peds_total,
            "other_t": self.other_t,
            "shi_t": self.shi_t,
            "unknown_age_t": self.unknown_age_t,
            "unknown_age_total": self.unknown_age_total,
            "adults_shi_t": self.adults_shi_t,
            "adults_other_t": self.adults_other_t,
            "peds_shi_t": self.peds_shi_t,
            "peds_other_t": self.peds_other_t,
            "by_month": {k: dict(self.by_month[k]) for k in sorted(self.by_month)},
        }


@dataclass
class GroupAgg:
    key: str
    name: str
    subtitle: str
    rows: list[IpaAgg] = field(default_factory=list)

    def totals(self) -> IpaAgg:
        t = IpaAgg(ipa="Total", sheet="")
        for r in self.rows:
            t.transferred       += r.transferred
            t.total             += r.total
            t.adults_t          += r.adults_t
            t.adults_total      += r.adults_total
            t.peds_t            += r.peds_t
            t.peds_total        += r.peds_total
            t.other_t           += r.other_t
            t.shi_t             += r.shi_t
            t.unknown_age_t     += r.unknown_age_t
            t.unknown_age_total += r.unknown_age_total
            t.adults_shi_t      += r.adults_shi_t
            t.adults_other_t    += r.adults_other_t
            t.peds_shi_t        += r.peds_shi_t
            t.peds_other_t      += r.peds_other_t
            for month, bucket in r.by_month.items():
                dst = t.by_month.setdefault(month, _new_month_bucket())
                _merge_month_bucket(dst, bucket)
        return t


# --- The main pass over one sheet ------------------------------------------

@dataclass
class SheetWarning:
    sheet: str
    kind: str
    detail: str


def aggregate_sheet(
    wb: openpyxl.Workbook,
    sheet_name: str,
    period_start: _dt.date | None,
    period_end: _dt.date | None,
    warnings: list[SheetWarning] | None = None,
) -> IpaAgg | None:
    """Return an IpaAgg for sheet_name, or None if the sheet is empty /
    has no recognizable header."""
    ws = wb[sheet_name]
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        return None

    cols = detect_columns(header)
    if cols is None:
        if warnings is not None:
            warnings.append(SheetWarning(sheet_name, "no_header",
                f"could not detect required columns; header={header[:8]}"))
        return None

    agg = IpaAgg(ipa=pretty_ipa_name(sheet_name), sheet=sheet_name)

    c_date = cols["date"]
    c_dob = cols["dob"]
    c_to = cols["transfer_to"]
    c_status = cols.get("transfer_status")
    c_eff = cols.get("effective_date")
    c_name = cols.get("patient_name")

    # All column indices we'll touch (used for blank-row detection)
    touched_cols = [i for i in (c_date, c_dob, c_to, c_status, c_eff, c_name)
                    if i is not None]
    max_idx = max(touched_cols)

    for row in rows_iter:
        if row is None:
            continue
        if len(row) <= max_idx:
            # Pad if openpyxl returned a short tuple
            row = tuple(row) + (None,) * (max_idx + 1 - len(row))

        # Non-blank row test: any value in the touched columns counts
        def _has(v):
            return v is not None and not (isinstance(v, str) and v.strip() == "")

        if not any(_has(row[i]) for i in touched_cols):
            continue

        date_v = coerce_date(row[c_date])

        # Period filter (when applied). Rows without a parseable DATE are
        # excluded from the period only when a period is in force.
        if period_start is not None and period_end is not None:
            if date_v is None:
                continue
            if date_v < period_start or date_v > period_end:
                continue

        # Resolve the monthly bucket (None if the row has no parseable DATE
        # and we're not enforcing a period — such rows still affect rollups
        # but won't show up under a month filter).
        month_key = date_v.strftime("%Y-%m") if date_v is not None else None
        bucket = (agg.by_month.setdefault(month_key, _new_month_bucket())
                  if month_key is not None else None)

        # Counted in Total
        agg.total += 1
        if bucket is not None:
            bucket["total"] += 1

        to_v = row[c_to]
        transferred = _has(to_v) and not is_excluded_status(
            row[c_status] if c_status is not None else None)

        shi = is_shi(to_v) if transferred else False
        if transferred:
            agg.transferred += 1
            if bucket is not None:
                bucket["transferred"] += 1
            if shi:
                agg.shi_t += 1
                if bucket is not None:
                    bucket["shi_t"] += 1
            else:
                agg.other_t += 1
                if bucket is not None:
                    bucket["other_t"] += 1

        # Age split applies to BOTH Total and Transferred (2024 convention:
        # adults_total + peds_total = total).
        eff_v = coerce_date(row[c_eff]) if c_eff is not None else None
        ref = eff_v or date_v
        dob = coerce_date(row[c_dob])
        a = age_on(dob, ref)
        if a is None:
            # Bucket unknowns into Adults to keep adults_total+peds_total==total.
            agg.adults_total += 1
            if bucket is not None:
                bucket["adults_total"] += 1
            if transferred:
                agg.adults_t += 1
                if bucket is not None:
                    bucket["adults_t"] += 1
                if shi:
                    agg.adults_shi_t += 1
                    if bucket is not None:
                        bucket["adults_shi_t"] += 1
                else:
                    agg.adults_other_t += 1
                    if bucket is not None:
                        bucket["adults_other_t"] += 1
            agg.unknown_age_total += 1
            if bucket is not None:
                bucket["unknown_age_total"] += 1
            if transferred:
                agg.unknown_age_t += 1
                if bucket is not None:
                    bucket["unknown_age_t"] += 1
            if (warnings is not None and len(warnings) < MAX_WARNINGS
                    and dob is None and row[c_dob] is not None):
                warnings.append(SheetWarning(
                    sheet_name, "unparseable_dob",
                    f"row date={date_v} dob={row[c_dob]!r}"))
        else:
            is_peds = a < 21
            if is_peds:
                agg.peds_total += 1
                if bucket is not None:
                    bucket["peds_total"] += 1
                if transferred:
                    agg.peds_t += 1
                    if bucket is not None:
                        bucket["peds_t"] += 1
                    if shi:
                        agg.peds_shi_t += 1
                        if bucket is not None:
                            bucket["peds_shi_t"] += 1
                    else:
                        agg.peds_other_t += 1
                        if bucket is not None:
                            bucket["peds_other_t"] += 1
            else:
                agg.adults_total += 1
                if bucket is not None:
                    bucket["adults_total"] += 1
                if transferred:
                    agg.adults_t += 1
                    if bucket is not None:
                        bucket["adults_t"] += 1
                    if shi:
                        agg.adults_shi_t += 1
                        if bucket is not None:
                            bucket["adults_shi_t"] += 1
                    else:
                        agg.adults_other_t += 1
                        if bucket is not None:
                            bucket["adults_other_t"] += 1

    return agg


MAX_WARNINGS = 200


def aggregate_workbook(
    xlsx_path: str,
    period_start: _dt.date | None,
    period_end: _dt.date | None,
) -> tuple[list[GroupAgg], list[SheetWarning], list[str]]:
    """Return (groups, warnings, unclassified_sheets)."""
    import warnings as _w
    _w.filterwarnings("ignore", category=UserWarning, module="openpyxl")

    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    groups: dict[str, GroupAgg] = {}
    warnings: list[SheetWarning] = []
    unclassified: list[str] = []

    for sheet_name in wb.sheetnames:
        cls = classify_sheet(sheet_name)
        if cls is None:
            if sheet_name.strip().upper() not in EXCLUDED_SHEETS:
                unclassified.append(sheet_name)
            continue
        key, display, subtitle = cls
        if key not in groups:
            groups[key] = GroupAgg(key=key, name=display, subtitle=subtitle)
        agg = aggregate_sheet(wb, sheet_name, period_start, period_end, warnings)
        if agg is None:
            # Empty sheet or undetectable header. Emit a zero-row entry so the
            # dashboard still lists the IPA.
            agg = IpaAgg(ipa=pretty_ipa_name(sheet_name), sheet=sheet_name)
        groups[key].rows.append(agg)

    # Order groups: Hanna, Samala, La Mirada, Benny B, Sakhai
    order = ["hanna", "samala", "la_mirada", "benny_b", "sakhai"]
    ordered = [groups[k] for k in order if k in groups]

    # Sort IPAs inside each group by total desc, then by name
    for g in ordered:
        g.rows.sort(key=lambda r: (-r.total, r.ipa))

    return ordered, warnings, unclassified
